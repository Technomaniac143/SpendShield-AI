import csv
import hashlib
import io
import logging
import re
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import PurePath
from typing import Iterable

from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.database import SessionLocal
from app.integrations.storage import ObjectStorage
from app.models import (
    GoodsReceipt, GoodsReceiptItem, IngestionError, IngestionFile, IngestionJob, IngestionStatus,
    Inventory, InventoryMovement, Invoice, InvoiceItem, Payment, Product, PurchaseOrder,
    PurchaseOrderItem, Supplier,
)
from app.services.entity_resolution import DeterministicEntityResolver, normalize

logger = logging.getLogger(__name__)
MAX_INGESTION_BYTES = 100 * 1024 * 1024
SUPPORTED_ENTITY_TYPES = {
    "suppliers", "products", "purchase_orders", "purchase_order_items", "goods_receipts",
    "goods_receipt_items", "invoices", "invoice_items", "payments", "inventory", "inventory_movements",
}
ERROR_CODES = {"INVALID_FILE", "FILE_TOO_LARGE", "INVALID_CSV", "INVALID_EXCEL", "MISSING_COLUMN", "INVALID_VALUE", "INVALID_DATE", "INVALID_DECIMAL", "INVALID_CURRENCY", "UNKNOWN_SUPPLIER", "UNKNOWN_PRODUCT", "UNKNOWN_PO", "UNKNOWN_GRN", "DUPLICATE_RECORD", "INVALID_REFERENCE", "ENTITY_REQUIRES_REVIEW", "INGESTION_CANCELLED", "INGESTION_FAILED"}


class IngestionValidationError(ValueError):
    def __init__(self, code: str, message: str, field: str | None = None, raw_value: str | None = None):
        super().__init__(message)
        self.code, self.field, self.raw_value = code, field, raw_value


def _date(value: str, field: str) -> date:
    try:
        return date.fromisoformat(value.strip())
    except (ValueError, AttributeError):
        raise IngestionValidationError("INVALID_DATE", "expected ISO date YYYY-MM-DD", field, value) from None


def _decimal(value: str, field: str, nonnegative: bool = True) -> Decimal:
    try:
        parsed = Decimal(value.strip())
    except (InvalidOperation, AttributeError):
        raise IngestionValidationError("INVALID_DECIMAL", "expected a decimal number", field, value) from None
    if not parsed.is_finite() or (nonnegative and parsed < 0):
        raise IngestionValidationError("INVALID_VALUE", "value must be finite and nonnegative", field, value)
    return parsed


def _currency(value: str) -> str:
    currency = value.strip().upper()
    if not re.fullmatch(r"[A-Z]{3}", currency):
        raise IngestionValidationError("INVALID_CURRENCY", "currency must be a three-letter ISO code", "currency", value)
    return currency


def _required(row: dict[str, str], fields: set[str]) -> None:
    missing = sorted(field for field in fields if not row.get(field, "").strip())
    if missing:
        raise IngestionValidationError("MISSING_COLUMN", f"missing required value: {missing[0]}", missing[0])


def _bytes_stream(content: bytes | object, job_type: str) -> Iterable[dict[str, str]]:
    if not content:
        raise IngestionValidationError("INVALID_FILE", "file is empty")
    if job_type == "CSV":
        try:
            binary = io.BytesIO(content) if isinstance(content, bytes) else content
            text = io.TextIOWrapper(binary, encoding="utf-8-sig", newline="")
            reader = csv.DictReader(text, strict=True)
            if not reader.fieldnames:
                raise IngestionValidationError("INVALID_CSV", "CSV header is required")
            if any(not name or not name.strip() for name in reader.fieldnames):
                raise IngestionValidationError("INVALID_CSV", "CSV contains an empty header")
            for row in reader:
                if None in row:
                    raise IngestionValidationError("INVALID_CSV", "CSV row has more values than headers")
                yield {key.strip(): (value or "") for key, value in row.items()}
        except csv.Error:
            raise IngestionValidationError("INVALID_CSV", "CSV contains malformed quoting or delimiters") from None
        except UnicodeDecodeError:
            raise IngestionValidationError("INVALID_CSV", "CSV must be UTF-8 encoded") from None
        return
    if not content.startswith(b"PK"):
        raise IngestionValidationError("INVALID_EXCEL", "file is not a valid XLSX archive")
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            raise IngestionValidationError("INVALID_EXCEL", "spreadsheet header is required")
        names = [str(value).strip() if value is not None else "" for value in headers]
        if any(not name for name in names):
            raise IngestionValidationError("INVALID_EXCEL", "spreadsheet contains an empty header")
        for values in rows:
            yield {name: "" if value is None else str(value) for name, value in zip(names, values)}
        workbook.close()
    except IngestionValidationError:
        raise
    except Exception as exc:
        raise IngestionValidationError("INVALID_EXCEL", "unable to read XLSX file") from exc


def _supplier(db: Session, tenant_id: str, value: str) -> Supplier:
    record = DeterministicEntityResolver(db).resolve_supplier(tenant_id, value)
    if record is None:
        raise IngestionValidationError("UNKNOWN_SUPPLIER", "supplier could not be resolved", "supplier", value)
    return record


def _product(db: Session, tenant_id: str, value: str) -> Product:
    record = DeterministicEntityResolver(db).resolve_product(tenant_id, value)
    if record is None:
        raise IngestionValidationError("UNKNOWN_PRODUCT", "product could not be resolved by SKU or name", "product", value)
    return record


def _record_row(db: Session, job: IngestionJob, row: dict[str, str]) -> None:
    entity = job.entity_type
    tenant_id = job.tenant_id
    _required(row, {
        "suppliers": {"name"}, "products": {"sku", "name"}, "purchase_orders": {"po_number", "supplier", "order_date", "currency"},
        "purchase_order_items": {"po_number", "product", "description", "quantity", "unit_price"}, "goods_receipts": {"grn_number", "po_number", "supplier", "receipt_date"},
        "goods_receipt_items": {"grn_number", "product", "quantity_received", "accepted_quantity", "rejected_quantity"}, "invoices": {"invoice_number", "supplier", "invoice_date", "currency"},
        "invoice_items": {"invoice_number", "product", "description", "quantity", "unit_price"}, "payments": {"invoice_number", "supplier", "amount", "currency", "payment_reference"},
        "inventory": {"warehouse", "product", "quantity", "unit_cost"}, "inventory_movements": {"warehouse", "product", "quantity", "movement_type", "movement_date"},
    }[entity])
    if entity == "suppliers":
        name = row["name"].strip()
        existing = db.scalar(select(Supplier).where(Supplier.tenant_id == tenant_id, Supplier.normalized_name == normalize(name)))
        if existing:
            return
        db.add(Supplier(tenant_id=tenant_id, name=name, normalized_name=normalize(name), tax_id=row.get("tax_id") or None, registration_id=row.get("registration_id") or None, address=row.get("address") or None, country=(row.get("country") or "").upper() or None, status=(row.get("status") or "ACTIVE").upper()))
    elif entity == "products":
        sku = row["sku"].strip()
        if db.scalar(select(Product).where(Product.tenant_id == tenant_id, Product.sku == sku)):
            return
        db.add(Product(tenant_id=tenant_id, sku=sku, name=row["name"].strip(), normalized_name=normalize(row["name"]), unit_of_measure=(row.get("unit_of_measure") or "UNIT").upper()))
    elif entity == "purchase_orders":
        supplier = _supplier(db, tenant_id, row["supplier"])
        if db.scalar(select(PurchaseOrder).where(PurchaseOrder.tenant_id == tenant_id, PurchaseOrder.po_number == row["po_number"].strip())):
            return
        db.add(PurchaseOrder(tenant_id=tenant_id, po_number=row["po_number"].strip(), supplier_id=supplier.id, order_date=_date(row["order_date"], "order_date"), currency=_currency(row["currency"]), status=(row.get("status") or "OPEN").upper(), total_amount=_decimal(row.get("total_amount", "0"), "total_amount")))
    elif entity == "purchase_order_items":
        po = db.scalar(select(PurchaseOrder).where(PurchaseOrder.tenant_id == tenant_id, PurchaseOrder.po_number == row["po_number"].strip()))
        if po is None: raise IngestionValidationError("UNKNOWN_PO", "purchase order could not be resolved", "po_number", row["po_number"])
        product = _product(db, tenant_id, row["product"])
        quantity, price = _decimal(row["quantity"], "quantity"), _decimal(row["unit_price"], "unit_price")
        tax, discount = _decimal(row.get("tax", "0"), "tax"), _decimal(row.get("discount", "0"), "discount")
        db.add(PurchaseOrderItem(tenant_id=tenant_id, purchase_order_id=po.id, product_id=product.id, description=row["description"], quantity=quantity, unit_price=price, tax=tax, discount=discount, line_total=quantity * price + tax - discount))
    elif entity == "goods_receipts":
        po = db.scalar(select(PurchaseOrder).where(PurchaseOrder.tenant_id == tenant_id, PurchaseOrder.po_number == row["po_number"].strip()))
        supplier = _supplier(db, tenant_id, row["supplier"])
        if po is None: raise IngestionValidationError("UNKNOWN_PO", "purchase order could not be resolved", "po_number", row["po_number"])
        if po.supplier_id != supplier.id: raise IngestionValidationError("INVALID_REFERENCE", "supplier does not match purchase order", "supplier", row["supplier"])
        db.add(GoodsReceipt(tenant_id=tenant_id, grn_number=row["grn_number"], po_id=po.id, supplier_id=supplier.id, receipt_date=_date(row["receipt_date"], "receipt_date"), status=(row.get("status") or "RECEIVED").upper()))
    elif entity == "goods_receipt_items":
        grn = db.scalar(select(GoodsReceipt).where(GoodsReceipt.tenant_id == tenant_id, GoodsReceipt.grn_number == row["grn_number"].strip()))
        if grn is None: raise IngestionValidationError("UNKNOWN_GRN", "goods receipt could not be resolved", "grn_number", row["grn_number"])
        product = _product(db, tenant_id, row["product"])
        received, accepted, rejected = (_decimal(row[key], key) for key in ("quantity_received", "accepted_quantity", "rejected_quantity"))
        if accepted + rejected > received: raise IngestionValidationError("INVALID_VALUE", "accepted and rejected exceed received quantity", "accepted_quantity")
        db.add(GoodsReceiptItem(tenant_id=tenant_id, goods_receipt_id=grn.id, product_id=product.id, quantity_received=received, accepted_quantity=accepted, rejected_quantity=rejected))
    elif entity == "invoices":
        supplier = _supplier(db, tenant_id, row["supplier"])
        if db.scalar(select(Invoice).where(Invoice.tenant_id == tenant_id, Invoice.invoice_number == row["invoice_number"].strip(), Invoice.supplier_id == supplier.id)): return
        po = None
        if row.get("po_number"):
            po = db.scalar(select(PurchaseOrder).where(PurchaseOrder.tenant_id == tenant_id, PurchaseOrder.po_number == row["po_number"].strip()))
            if po is None: raise IngestionValidationError("UNKNOWN_PO", "purchase order could not be resolved", "po_number", row["po_number"])
        db.add(Invoice(tenant_id=tenant_id, invoice_number=row["invoice_number"].strip(), supplier_id=supplier.id, po_id=po.id if po else None, invoice_date=_date(row["invoice_date"], "invoice_date"), currency=_currency(row["currency"]), subtotal=_decimal(row.get("subtotal", "0"), "subtotal"), tax=_decimal(row.get("tax", "0"), "tax"), discount=_decimal(row.get("discount", "0"), "discount"), total_amount=_decimal(row.get("total_amount", "0"), "total_amount"), status=(row.get("status") or "RECEIVED").upper(), document_hash=row.get("document_hash") or None))
    elif entity == "invoice_items":
        invoice = db.scalar(select(Invoice).join(Supplier, Supplier.id == Invoice.supplier_id).where(Invoice.tenant_id == tenant_id, Invoice.invoice_number == row["invoice_number"].strip()))
        if invoice is None: raise IngestionValidationError("INVALID_REFERENCE", "invoice could not be resolved", "invoice_number", row["invoice_number"])
        product = _product(db, tenant_id, row["product"])
        quantity, price = _decimal(row["quantity"], "quantity"), _decimal(row["unit_price"], "unit_price")
        tax, discount = _decimal(row.get("tax", "0"), "tax"), _decimal(row.get("discount", "0"), "discount")
        db.add(InvoiceItem(tenant_id=tenant_id, invoice_id=invoice.id, product_id=product.id, description=row["description"], quantity=quantity, unit_price=price, tax=tax, discount=discount, line_total=quantity * price + tax - discount))
    elif entity == "payments":
        supplier = _supplier(db, tenant_id, row["supplier"])
        invoice = db.scalar(select(Invoice).where(Invoice.tenant_id == tenant_id, Invoice.invoice_number == row["invoice_number"], Invoice.supplier_id == supplier.id))
        if invoice is None: raise IngestionValidationError("INVALID_REFERENCE", "invoice could not be resolved", "invoice_number", row["invoice_number"])
        db.add(Payment(tenant_id=tenant_id, invoice_id=invoice.id, amount=_decimal(row["amount"], "amount"), currency=_currency(row["currency"]), payment_date=_date(row["payment_date"], "payment_date") if row.get("payment_date") else None, status=(row.get("status") or "PENDING").upper(), payment_reference=row["payment_reference"].strip()))
    elif entity == "inventory":
        product = _product(db, tenant_id, row["product"])
        quantity, unit_cost = _decimal(row["quantity"], "quantity"), _decimal(row["unit_cost"], "unit_cost")
        if db.scalar(select(Inventory).where(Inventory.tenant_id == tenant_id, Inventory.warehouse == row["warehouse"], Inventory.product_id == product.id)): return
        db.add(Inventory(tenant_id=tenant_id, warehouse=row["warehouse"].strip(), product_id=product.id, quantity=quantity, unit_cost=unit_cost, inventory_value=quantity * unit_cost, last_movement=_date(row["last_movement"], "last_movement") if row.get("last_movement") else None))
    elif entity == "inventory_movements":
        product = _product(db, tenant_id, row["product"])
        inventory = db.scalar(select(Inventory).where(Inventory.tenant_id == tenant_id, Inventory.warehouse == row["warehouse"], Inventory.product_id == product.id))
        if inventory is None: raise IngestionValidationError("INVALID_REFERENCE", "inventory could not be resolved", "warehouse")
        movement = row["movement_type"].upper(); quantity = _decimal(row["quantity"], "quantity")
        if movement in {"OUT", "ISSUE", "CONSUMPTION"} and quantity > inventory.quantity: raise IngestionValidationError("INVALID_VALUE", "movement exceeds available inventory", "quantity")
        inventory.quantity = inventory.quantity - quantity if movement in {"OUT", "ISSUE", "CONSUMPTION"} else inventory.quantity + quantity
        inventory.inventory_value = inventory.quantity * inventory.unit_cost
        inventory.last_movement = _date(row["movement_date"], "movement_date")
        db.add(InventoryMovement(tenant_id=tenant_id, inventory_id=inventory.id, quantity=quantity, movement_type=movement, movement_date=inventory.last_movement, reference=row.get("reference") or None))


def trigger_post_ingestion_analytics(job_id: str) -> None:
    logger.info("post-ingestion analytics hook queued", extra={"job_id": job_id})


def process_ingestion_job(job_id: str, storage: ObjectStorage) -> None:
    with SessionLocal() as db:
        job = db.get(IngestionJob, job_id)
        if job is None or job.status in {IngestionStatus.CANCELLED, IngestionStatus.COMPLETED, IngestionStatus.PARTIAL}: return
        file_record = db.scalar(select(IngestionFile).where(IngestionFile.job_id == job.id, IngestionFile.tenant_id == job.tenant_id))
        if file_record is None: job.status, job.last_error = IngestionStatus.FAILED, "ingestion file is missing"; db.commit(); return
        job.status, job.started_at = IngestionStatus.PROCESSING, datetime.now(timezone.utc); db.commit()
        try:
            content = storage.get_stream(file_record.object_key) if job.job_type == "CSV" and hasattr(storage, "get_stream") else storage.get(file_record.object_key)
            for row_number, row in enumerate(_bytes_stream(content, job.job_type), start=2):
                if job.status == IngestionStatus.CANCELLED: break
                job.total_rows = max(job.total_rows, row_number - 1)
                try:
                    _record_row(db, job, row)
                    db.commit(); job.successful_rows += 1
                except IngestionValidationError as exc:
                    db.rollback(); db.add(IngestionError(job_id=job.id, tenant_id=job.tenant_id, row_number=row_number, field=exc.field, error_code=exc.code, message=str(exc), raw_value=exc.raw_value)); db.commit(); job.failed_rows += 1
                except IntegrityError:
                    db.rollback(); db.add(IngestionError(job_id=job.id, tenant_id=job.tenant_id, row_number=row_number, field=None, error_code="DUPLICATE_RECORD", message="record conflicts with an existing record", raw_value=None)); db.commit(); job.failed_rows += 1
                job.processed_rows += 1; job.total_rows = max(job.total_rows, job.processed_rows); db.commit()
            if job.status != IngestionStatus.CANCELLED:
                job.status = IngestionStatus.PARTIAL if job.failed_rows else IngestionStatus.COMPLETED
                job.completed_at = datetime.now(timezone.utc); db.commit(); trigger_post_ingestion_analytics(job.id)
        except IngestionValidationError as exc:
            job.status, job.last_error, job.completed_at = IngestionStatus.FAILED, str(exc), datetime.now(timezone.utc); db.commit()
        except Exception:
            logger.exception("ingestion job failed", extra={"job_id": job.id})
            job.status, job.last_error, job.completed_at = IngestionStatus.FAILED, "ingestion processing failed", datetime.now(timezone.utc); db.commit()


def create_job(db: Session, tenant_id: str, actor: str, entity_type: str, job_type: str, filename: str, content: bytes, storage: ObjectStorage, idempotency_key: str | None):
    if entity_type not in SUPPORTED_ENTITY_TYPES: raise ValueError("unsupported entity_type")
    digest = hashlib.sha256(content).hexdigest()
    if idempotency_key:
        existing = db.scalar(select(IngestionJob).where(IngestionJob.tenant_id == tenant_id, IngestionJob.idempotency_key == idempotency_key))
        if existing: return existing
    existing_file = db.scalar(select(IngestionFile).where(IngestionFile.tenant_id == tenant_id, IngestionFile.sha256 == digest))
    if existing_file: return db.get(IngestionJob, existing_file.job_id)
    safe_name = re.sub(r"[^A-Za-z0-9._-]", "_", PurePath(filename).name.replace("..", "_")) or "upload"
    job = IngestionJob(tenant_id=tenant_id, job_type=job_type, entity_type=entity_type, status=IngestionStatus.QUEUED, source_filename=safe_name[:255], file_hash=digest, created_by=actor, idempotency_key=idempotency_key)
    db.add(job); db.flush()
    object_key = f"{tenant_id}/ingestion/{job.id}/{hashlib.sha256(safe_name.encode()).hexdigest()[:16]}-{safe_name}"
    try:
        storage.put(object_key, content, "text/csv" if job_type == "CSV" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        db.add(IngestionFile(job_id=job.id, tenant_id=tenant_id, object_key=object_key, content_type="text/csv" if job_type == "CSV" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", size_bytes=len(content), sha256=digest)); db.commit(); return job
    except Exception:
        db.rollback()
        try: storage.delete(object_key)
        except Exception: logger.exception("failed to clean up ingestion object", extra={"object_key": object_key})
        raise